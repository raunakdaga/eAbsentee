import yagmail
from datetime import date
from keys import GMAIL_SENDER_ADDRESS, GMAIL_SENDER_PASSWORD
import os
import openpyxl
from openpyxl import load_workbook
import json
import sys

# Change current working directory, only needed for Atom
os.chdir(os.path.dirname(os.path.abspath(__file__)))
os.chdir('../reports')


def split_reports() -> None:
    pass


def email_report_daily() -> None:
    # print(os.path.dirname(os.path.realpath(__file__)))
    """Email the Excel spreadsheet to Senator Surovell and Mr. Rouvelas. """
    today_date: str = date.today().strftime("%m-%d-%y")
    report_path = f'{today_date}.xlsx'
    report: openpyxl.workbook.Workbook = load_workbook(filename=report_path)
    worksheet: openpyxl.worksheet.worksheet.Worksheet = report.active
    if worksheet['A2'].value:
        yagmail.SMTP(GMAIL_SENDER_ADDRESS, GMAIL_SENDER_PASSWORD).send(
            to=['raunakdaga@gmail.com', 'lerouvelas@gmail.com'],
            subject=f'Daily Absentee Ballot Application Report - {today_date}',
            contents=f'Please find attached the daily report of absentee ' +
            f'ballot applications for {today_date}.',
            attachments=report_path
        )


def email_report(file_name, emails):
    today_date: str = date.today().strftime("%m-%d-%y")
    report: openpyxl.workbook.Workbook = load_workbook(filename='surovell.xlsx')
    worksheet: openpyxl.worksheet.worksheet.Worksheet = report.active
    if worksheet['B' + str(worksheet.max_row)].value.split()[0] == today_date:
        yagmail.SMTP(GMAIL_SENDER_ADDRESS, GMAIL_SENDER_PASSWORD).send(
            # to='raunakdaga@gmail.com',
            to=emails,
            subject=f'Daily Report - eAbsentee Applications',
            contents=f'New absentee ballot applications were submitted ' +
            'yesterday using your eAbsentee.org campaign link. Attached ' +
            'is a report on new and previously-submitted applications.',
            attachments=file_name
        )


def email_all_groups():
    print(os.getcwd())
    with open('../static/groups.json') as file:
        groups_json = json.load(file)
        groups = groups_json.keys()
        for group in groups:
            if os.path.isfile(group + '.xlsx'):
                try:
                    email_report(group + '.xlsx', groups_json[group]['email'].split())
                except:
                    print('Oops, an error occurred with the group ' + group, file=sys.stderr)


email_report_daily()
# email_all_groups()


# today_date: str = date.today().strftime("%m-%d-%y")
# report: openpyxl.workbook.Workbook = load_workbook(filename='surovell.xlsx')
# worksheet: openpyxl.worksheet.worksheet.Worksheet = report.active
# print(worksheet['B' + str(worksheet.max_row)].value.split()[0] == today_date)
